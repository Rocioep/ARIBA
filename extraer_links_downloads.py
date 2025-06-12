# extraer_links_downloads.py
"""
Extrae links reales de archivos Excel en downloads
"""

import openpyxl
from pathlib import Path

def extraer_links_desde_downloads():
    """Extraer hipervínculos desde la carpeta downloads"""
    
    # Ruta específica donde están tus archivos
    downloads_dir = Path("alfamine_monitor_v1/data/downloads")
    
    if not downloads_dir.exists():
        print(f"❌ No existe la carpeta: {downloads_dir}")
        return
    
    # Buscar archivos data
    archivos = list(downloads_dir.glob("data*"))
    
    if not archivos:
        print("❌ No se encontraron archivos data en downloads")
        return
    
    print("📁 Archivos encontrados:")
    for i, archivo in enumerate(archivos, 1):
        size_kb = archivo.stat().st_size / 1024
        print(f"  {i}. {archivo.name} ({size_kb:.0f} KB)")
    
    # Usar data(4) si existe, sino el más reciente
    archivo_elegido = None
    for archivo in archivos:
        if "data(4)" in archivo.name:
            archivo_elegido = archivo
            break
    
    if not archivo_elegido:
        archivo_elegido = max(archivos, key=lambda x: x.stat().st_mtime)
    
    print(f"\n🎯 Procesando: {archivo_elegido.name}")
    
    try:
        # Abrir con openpyxl
        workbook = openpyxl.load_workbook(archivo_elegido)
        sheet = workbook.active
        
        print(f"📊 Hoja: {sheet.title}")
        print(f"📋 Filas: {sheet.max_row}, Columnas: {sheet.max_column}")
        
        links_encontrados = []
        
        print("\n🔍 Buscando hipervínculos desde fila 3...")
        
        # Desde fila 3 hacia abajo
        for row in range(3, min(sheet.max_row + 1, 50)):  # Limitar a 50 para prueba
            cell = sheet[f'A{row}']
            
            if cell.hyperlink:
                url_real = cell.hyperlink.target
                texto = str(cell.value) if cell.value else ""
                
                print(f"✅ Fila {row}: {texto[:40]}...")
                print(f"   Link: {url_real[:80]}...")
                
                links_encontrados.append({
                    'fila': row,
                    'texto': texto,
                    'url': url_real
                })
            else:
                # Mostrar algunas filas sin link para debug
                if row <= 10:
                    texto = str(cell.value) if cell.value else ""
                    if texto.strip():
                        print(f"❌ Fila {row}: Sin hipervínculo - '{texto[:30]}...'")
        
        workbook.close()
        
        print(f"\n🎉 TOTAL ENCONTRADOS: {len(links_encontrados)} hipervínculos")
        
        if links_encontrados:
            # Guardar resultados
            with open('links_extraidos.txt', 'w', encoding='utf-8') as f:
                f.write(f"LINKS EXTRAÍDOS DE {archivo_elegido.name}\n")
                f.write("=" * 60 + "\n\n")
                
                for i, link in enumerate(links_encontrados, 1):
                    f.write(f"{i}. Fila {link['fila']}:\n")
                    f.write(f"   Texto: {link['texto']}\n")
                    f.write(f"   URL: {link['url']}\n")
                    f.write("-" * 40 + "\n")
            
            print("💾 Links guardados en: links_extraidos.txt")
            
            # Mostrar primer link como ejemplo
            if links_encontrados:
                primer_link = links_encontrados[0]
                print(f"\n📋 EJEMPLO - Primer link encontrado:")
                print(f"   Fila: {primer_link['fila']}")
                print(f"   Texto: {primer_link['texto'][:50]}...")
                print(f"   URL: {primer_link['url']}")
                print("\n💡 Copia esta URL y pégala en una nueva pestaña para probar")
        
        else:
            print("\n🤔 No se encontraron hipervínculos. Posibles causas:")
            print("   • Los links están en otra columna")
            print("   • El archivo no tiene hipervínculos reales")
            print("   • Necesita otro método de extracción")
        
        return links_encontrados
        
    except Exception as e:
        print(f"❌ Error procesando archivo: {e}")
        return []

if __name__ == "__main__":
    print("🔍 EXTRACTOR DE LINKS DESDE DOWNLOADS")
    print("=" * 50)
    extraer_links_desde_downloads()