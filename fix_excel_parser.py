# fix_excel_parser.py
"""
FIX ESPECÍFICO para parsear Excel principal
Links en Columna A desde fila 3 hacia abajo
VERSIÓN COMPLETA Y FUNCIONAL
"""

import openpyxl
import pandas as pd
from pathlib import Path
import re
import json
from datetime import datetime

def parse_excel_principal_fixed(excel_path):
    """Analizar Excel principal CORREGIDO - Maneja .xls (HTML) y .xlsx"""
    try:
        print(f"📖 Analizando Excel principal: {excel_path.name}")
        
        # Verificar formato del archivo
        if excel_path.suffix.lower() == '.xls':
            # Archivo .xls - probablemente HTML con extensión .xls
            print("📄 Formato .xls detectado - usando método HTML...")
            return parse_xls_with_beautifulsoup(excel_path)
        else:
            # Archivo .xlsx - usar openpyxl
            print("📄 Formato .xlsx detectado - usando openpyxl...")
            return parse_xlsx_with_openpyxl(excel_path)
            
    except Exception as e:
        print(f"❌ Error analizando Excel: {e}")
        return []

def parse_xlsx_with_openpyxl(excel_path):
    """Parsear .xlsx con openpyxl (para hipervínculos reales)"""
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        
        print(f"📊 Hoja: {sheet.title}")
        print(f"📋 Filas: {sheet.max_row}, Columnas: {sheet.max_column}")
        
        licitaciones = []
        
        print("🔍 Buscando hipervínculos desde fila 3...")
        
        for row_num in range(3, sheet.max_row + 1):
            try:
                cell = sheet[f'A{row_num}']
                
                # Verificar si la celda tiene hipervínculo
                if cell.hyperlink:
                    url_real = cell.hyperlink.target
                    texto_visible = str(cell.value) if cell.value else ""
                    
                    # Obtener ID de la columna B
                    id_cell = sheet[f'B{row_num}']
                    licitacion_id = str(id_cell.value) if id_cell.value else f"Licitacion_{row_num}"
                    
                    # Verificar si es un link de Ariba válido
                    if is_ariba_link_fixed(url_real):
                        print(f"✅ Fila {row_num}: HIPERVÍNCULO ENCONTRADO")
                        print(f"   Texto: {texto_visible[:50]}...")
                        print(f"   URL: {url_real[:70]}...")
                        
                        licitaciones.append({
                            'index': row_num - 1,
                            'id': licitacion_id,
                            'titulo': texto_visible,
                            'link': url_real,
                            'fila_excel': row_num,
                            'tipo': 'xlsx_hyperlink'
                        })
            
            except Exception as e:
                print(f"⚠️ Error procesando fila {row_num}: {e}")
                continue
        
        workbook.close()
        
        print(f"✅ {len(licitaciones)} hipervínculos extraídos de archivo .xlsx")
        return licitaciones
        
    except Exception as e:
        print(f"❌ Error procesando .xlsx: {e}")
        return []

def parse_xls_with_beautifulsoup(excel_path):
    """Parsear archivo HTML con extensión .xls y extraer hipervínculos"""
    print("🌐 Archivo es HTML con extensión .xls - extrayendo hipervínculos reales...")
    
    try:
        # Leer archivo HTML como texto
        with open(excel_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # Verificar que es HTML
        if not html_content.strip().startswith('<'):
            print("❌ El archivo no parece ser HTML")
            return []
        
        # Parsear HTML con BeautifulSoup
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # Buscar todos los enlaces (tag <a>)
        links = soup.find_all('a', href=True)
        
        print(f"🔗 {len(links)} enlaces encontrados en el HTML")
        
        licitaciones = []
        
        for i, link in enumerate(links, 1):
            try:
                href = link.get('href')
                texto = link.get_text(strip=True)
                
                # Verificar si es un link de Ariba válido
                if href and is_ariba_link_fixed(href):
                    print(f"✅ Link {i}: {texto[:50]}...")
                    print(f"   URL: {href[:80]}...")
                    
                    licitaciones.append({
                        'index': i,
                        'id': f"Licitacion_{i}",
                        'titulo': texto,
                        'link': href,
                        'fila_excel': i + 2,  # Aproximado
                        'tipo': 'html_link'
                    })
                
                elif href and i <= 5:
                    # Mostrar links no válidos para debug (solo primeros 5)
                    print(f"❌ Link {i}: No es Ariba - {href[:40]}...")
                        
            except Exception as e:
                print(f"⚠️ Error procesando link {i}: {e}")
                continue
        
        print(f"🎉 {len(licitaciones)} hipervínculos de Ariba extraídos")
        
        return licitaciones
        
    except ImportError:
        print("❌ BeautifulSoup no está instalado")
        print("💡 Instala con: pip install beautifulsoup4")
        
        # Fallback: intentar con pandas (sin hipervínculos)
        print("🔄 Intentando método alternativo con pandas...")
        return parse_xls_with_pandas_fallback(excel_path)
        
    except Exception as e:
        print(f"❌ Error procesando HTML: {e}")
        return []

def parse_xls_with_pandas_fallback(excel_path):
    """Método de respaldo usando pandas para leer HTML"""
    try:
        print("📊 Usando pandas como respaldo para leer HTML...")
        
        # Intentar leer como HTML con pandas
        df_list = pd.read_html(str(excel_path))
        
        if not df_list:
            print("❌ No se pudieron extraer tablas del HTML")
            return []
        
        # Tomar la tabla más grande
        df = max(df_list, key=len)
        
        print(f"📊 Tabla HTML leída: {len(df)} filas, {len(df.columns)} columnas")
        print(f"📋 Columnas: {list(df.columns)}")
        
        licitaciones = []
        columna_a = df.iloc[:, 0]  # Primera columna
        
        print("🔍 Analizando contenido HTML con pandas...")
        print("⚠️ NOTA: pandas no puede extraer hipervínculos, solo texto")
        
        for index in range(2, min(len(df), 50)):  # Limitar para prueba
            try:
                link_value = str(columna_a.iloc[index]) if pd.notna(columna_a.iloc[index]) else ""
                
                if link_value.strip() and link_value != 'nan':
                    print(f"📋 Fila {index+1}: '{link_value[:50]}...'")
                    
                    # Para pandas, solo podemos buscar URLs en el texto
                    if is_ariba_link_fixed(link_value):
                        row = df.iloc[index]
                        licitacion_id = str(row.iloc[1]) if len(row) > 1 and pd.notna(row.iloc[1]) else f"Licitacion_{index+1}"
                        
                        licitaciones.append({
                            'index': index,
                            'id': licitacion_id,
                            'titulo': link_value,
                            'link': link_value,
                            'fila_excel': index + 1,
                            'tipo': 'pandas_fallback'
                        })
                        
                        print(f"✅ Link candidato encontrado en fila {index+1}")
                        
            except Exception as e:
                continue
        
        print(f"📊 Total procesado con pandas: {len(licitaciones)} elementos")
        return licitaciones
        
    except Exception as e:
        print(f"❌ Error con método de respaldo: {e}")
        return []

def is_ariba_link_fixed(text):
    """Verificar si un texto contiene un link de Ariba válido - MEJORADO"""
    if not isinstance(text, str) or not text.strip():
        return False
    
    # Patrones más específicos para Ariba
    ariba_patterns = [
        r'https?://.*ariba\.com.*',
        r'https?://service\.ariba\.com.*',
        r'https?://.*supplier.*\.ariba\.com.*',
        r'https?://portaldenegocios-codelco\.supplier.*\.ariba\.com.*',
        r'.*aw\?.*awh=.*',  # Patrón específico de Ariba
        r'.*Sourcing\.aw.*',  # Otro patrón típico
        r'.*\.ariba\.com.*',
        r'.*webjumper\?itemID=.*'  # Patrón específico de links de Ariba
    ]
    
    for pattern in ariba_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return True
    
    # También verificar si contiene palabras clave de Ariba
    ariba_keywords = [
        'ariba.com', 
        'sourcing.aw', 
        'awh=', 
        'dard=', 
        'supplier.ariba.com',
        'webjumper',
        'itemID=',
        'portaldenegocios-codelco'
    ]
    
    text_lower = text.lower()
    
    for keyword in ariba_keywords:
        if keyword in text_lower:
            return True
    
    return False

def save_licitaciones_to_files(licitaciones):
    """Guardar licitaciones en múltiples formatos"""
    if not licitaciones:
        print("❌ No hay licitaciones para guardar")
        return
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 1. Guardar en JSON (para el sistema automático)
    json_file = f'licitaciones_{len(licitaciones)}_{timestamp}.json'
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(licitaciones, f, indent=2, ensure_ascii=False)
    print(f"💾 {len(licitaciones)} licitaciones guardadas en JSON: {json_file}")
    
    # 2. Guardar en TXT (para revisión manual)
    txt_file = f'licitaciones_{len(licitaciones)}_{timestamp}.txt'
    with open(txt_file, 'w', encoding='utf-8') as f:
        f.write(f"LICITACIONES EXTRAÍDAS DE ARIBA\n")
        f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total: {len(licitaciones)} licitaciones\n")
        f.write("=" * 70 + "\n\n")
        
        for i, lic in enumerate(licitaciones, 1):
            f.write(f"{i:3d}. {lic['titulo'][:60]}...\n")
            f.write(f"     ID: {lic['id']}\n")
            f.write(f"     URL: {lic['link']}\n")
            f.write(f"     Fila Excel: {lic['fila_excel']}\n")
            f.write("-" * 50 + "\n")
    
    print(f"📄 Lista legible guardada en TXT: {txt_file}")
    
    # 3. Guardar archivo simple para sistema automático
    simple_file = 'licitaciones_para_procesar.json'
    with open(simple_file, 'w', encoding='utf-8') as f:
        json.dump(licitaciones, f, indent=2, ensure_ascii=False)
    print(f"🤖 Archivo para sistema automático: {simple_file}")
    
    return json_file, txt_file, simple_file

def test_excel_parsing():
    """Probar el parseo de Excel con archivos que empiecen con 'data'"""
    
    # Buscar en data/downloads (desde alfamine_monitor_v1)
    downloads_dir = Path("data/downloads")
    
    # Buscar archivos Excel que empiecen con "data"
    data_files = list(downloads_dir.glob("data*.xlsx")) + list(downloads_dir.glob("data*.xls"))
    
    if not data_files:
        print("❌ No se encontraron archivos Excel que empiecen con 'data'")
        
        # Buscar todos los Excel como fallback
        all_excel = list(downloads_dir.glob("*.xlsx")) + list(downloads_dir.glob("*.xls"))
        if all_excel:
            print("📊 Archivos Excel encontrados (fallback):")
            for i, file in enumerate(all_excel, 1):
                size_mb = file.stat().st_size / (1024 * 1024)
                print(f"  {i}. {file.name} ({size_mb:.1f} MB)")
        else:
            print("❌ No se encontraron archivos Excel en data/downloads")
        return []
    
    # Mostrar archivos "data" encontrados
    print("📊 Archivos 'data' encontrados:")
    for i, file in enumerate(data_files, 1):
        size_mb = file.stat().st_size / (1024 * 1024)
        print(f"  {i}. {file.name} ({size_mb:.1f} MB)")
    
    # Probar con el más reciente de los archivos "data"
    latest_data_file = max(data_files, key=lambda x: x.stat().st_mtime)
    print(f"\n🔍 Probando con archivo 'data' más reciente: {latest_data_file.name}")
    
    # Parsear
    licitaciones = parse_excel_principal_fixed(latest_data_file)
    
    print(f"\n📊 RESULTADO: {len(licitaciones)} licitaciones con links válidos")
    
    # Guardar resultados si hay licitaciones
    if licitaciones:
        print("\n💾 Guardando resultados...")
        files_created = save_licitaciones_to_files(licitaciones)
        
        print(f"\n🎯 RESUMEN:")
        print(f"   • {len(licitaciones)} licitaciones extraídas")
        print(f"   • Archivos creados: {len(files_created)}")
        print(f"   • Listo para sistema automático: licitaciones_para_procesar.json")
        
        # Mostrar muestra
        print(f"\n📋 Muestra de licitaciones (primeras 5):")
        for i, lic in enumerate(licitaciones[:5], 1):
            print(f"  {i}. {lic['titulo'][:50]}...")
            print(f"     URL: {lic['link'][:60]}...")
        
        if len(licitaciones) > 5:
            print(f"     ... y {len(licitaciones) - 5} licitaciones más")
    
    return licitaciones

if __name__ == "__main__":
    print("🔍 FIX EXCEL PARSER - EXTRACTOR DE LICITACIONES")
    print("=" * 60)
    print("📅 Fecha:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print()
    
    try:
        licitaciones = test_excel_parsing()
        
        if licitaciones:
            print(f"\n🎉 ¡ÉXITO! {len(licitaciones)} licitaciones listas para procesar")
            print("\n🚀 Próximo paso: Ejecutar sistema automático")
            print("   python sistema_iterativo_completo.py")
        else:
            print("\n❌ No se pudieron extraer licitaciones")
            print("💡 Verifica que los archivos existan y tengan el formato correcto")
            
    except KeyboardInterrupt:
        print("\n⏹️ Proceso interrumpido por el usuario")
    except Exception as e:
        print(f"\n💥 Error inesperado: {e}")
        import traceback
        traceback.print_exc()